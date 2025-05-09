import pdfplumber
import fitz  # PyMuPDF
import re
from django.shortcuts import render, redirect
from .models import (
    EarmarkedTransaction, ParsedTransaction, 
    Property, Tenant, ExpenseProfile, Unit, 
    Landlord, Lease , IncomeProfile
)
from django.shortcuts import get_object_or_404
from datetime import datetime
import PyPDF2
from io import BytesIO
import openpyxl
from django.http import HttpResponse
from django.forms import modelformset_factory
from django.forms import model_to_dict
from django.http import JsonResponse
from decimal import Decimal
from datetime import datetime
from django.urls import reverse
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.db.models import Q
from itertools import chain
from operator import attrgetter
from django.contrib.auth.views import LoginView
from django.contrib.auth import authenticate
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.contrib import messages
from datetime import date
from django.db.models import Sum, F, FloatField, ExpressionWrapper
from django.core.exceptions import ValidationError
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from .models import Unit, Property
from openpyxl import Workbook
from django.http import HttpResponse
from bookkeeping.utils.onedrive_utils import upload_to_onedrive

# Custom Login View
class CustomLoginView(LoginView):
    template_name = 'login.html'  # Path to your login template

    def form_valid(self, form):
        user = form.get_user()
        request = self.request  # Get the request object
        force_login = request.POST.get("force_login") == "1"  # Check if force_login is set

        print(f"[DEBUG] User attempting login: {user.username} (ID: {user.id})")
        print(f"[DEBUG] force_login flag: {force_login}")

        existing_sessions = Session.objects.filter(expire_date__gte=timezone.now())  # Get all active sessions
        print(f"[DEBUG] Found {existing_sessions.count()} active sessions.")

        user_has_session = False  # Flag to check if user has an existing session

        # Check if the user already has an active session
        for session in existing_sessions:
            data = session.get_decoded()
            if data.get('_auth_user_id') == str(user.id):  # Check if the user ID matches
                print(f"[DEBUG] Existing session found for user {user.username} (Session Key: {session.session_key})")
                if not force_login:
                    user_has_session = True  # Set the flag to True if user has an existing session
                    break
                else:
                    session.delete()  # Delete the existing session if force_login is set
                    print(f"[DEBUG] Deleted existing session for user {user.username} due to force_login.")

        if user_has_session:  # If the user has an existing session and force_login is not set
            print(f"[DEBUG] Login prevented for user {user.username}: already logged in.")
            form.add_error(None, "already_logged_in")  # key only, used in template
            return self.form_invalid(form)

        print(f"[DEBUG] Login allowed for user {user.username}")
        return super().form_valid(form)


# Dashboard: Displays Earmarked and Parsed Transactions
@login_required
def dashboard(request):
    earmarked = EarmarkedTransaction.objects.all().order_by('booking_no')  # Order by booking number
    parsed = ParsedTransaction.objects.all().order_by('booking_no')        # Order by booking number
    properties = Property.objects.all()
    units = Unit.objects.all()
    leases = Lease.objects.all()
    return render(request, 'bookkeeping/dashboard.html', {
        'earmarked': earmarked,
        'parsed': parsed,
        'properties': properties,
        'units': units,
        'leases': leases,
    })

# Helper function to serialize transactions for session storage
# This function converts date objects to ISO format strings for session storage.
def serialize_for_session(tx_list):
    serialized = []
    for tx in tx_list:
        tx_copy = tx.copy()
        if isinstance(tx_copy.get('date'), date):
            tx_copy['date'] = tx_copy['date'].isoformat()  # e.g. '2024-04-10'
        serialized.append(tx_copy)
    return serialized

# Upload Bank Statement 
def upload_bank_statement(request, property_id=None):
    property_obj = get_object_or_404(Property, id=property_id) # Get the property object

    if request.method == 'POST' and request.FILES.get('statement'): 
        statement = request.FILES['statement'] 
        duplicate_count = 0
        earmarked_transactions = []
        current_buchungsdatum = None

        # Extract IBAN from first page
        reader = PyPDF2.PdfReader(statement)
        first_page_text = reader.pages[0].extract_text()
        iban_match = re.search(r"IBAN:\s*([A-Z0-9\s]+)", first_page_text)

        if not iban_match:
            return render(request, 'bookkeeping/upload_statement.html', {
                'property': property_obj,
                'error': "\u26a0\ufe0f IBAN konnte im PDF nicht gefunden werden.",
            })

        extracted_iban = re.sub(r"\s+", "", iban_match.group(1))[:22] # Clean and limit to 22 characters
        landlord_ibans = [iban.replace(" ", "") for iban in property_obj.landlords.values_list('iban', flat=True) if iban] # Get all IBANs of landlords associated with the property

        if extracted_iban not in landlord_ibans:
            messages.error(request, f"\u274c Die IBAN {extracted_iban} stimmt nicht mit den IBANs der Vermieter \u00fcberein.") 
            return redirect(f"{reverse('property_detail', args=[property_id])}?tab=dashboard") # Redirect to property detail page with error message

        # Booking Number Counter Setup
        latest_bn = EarmarkedTransaction.objects.filter(
            property=property_obj
        ).exclude(booking_no__isnull=True).order_by('-booking_no').first() # Get the latest booking number for the property
        base_bn = 1
        if latest_bn and latest_bn.booking_no:
            try:
                base_bn = int(re.sub(r'\D', '', latest_bn.booking_no)) + 1 # Increment booking number
            except ValueError:
                pass
        bn_counter = base_bn

        def split_pdf_into_pages(pdf_file):
            pdf_reader = PyPDF2.PdfReader(pdf_file) # Read the PDF file
            pages = []
            for page_num in range(len(pdf_reader.pages)):
                writer = PyPDF2.PdfWriter() # Create a PDF writer object
                writer.add_page(pdf_reader.pages[page_num]) # Add each page to the writer
                stream = BytesIO() # Create a stream to hold the page
                writer.write(stream) # Write the page to the stream
                stream.seek(0)   # Reset the stream position to the beginning
                pages.append(stream)  # Append the stream to the pages list
            return pages

        def extract_text_with_fallback(page_stream):
            try:
                with pdfplumber.open(page_stream) as pdf:
                    return pdf.pages[0].extract_text()  # Extract text using pdfplumber
            except Exception:
                page_stream.seek(0)
                with fitz.open(stream=page_stream.read(), filetype="pdf") as pdf:
                    return pdf[0].get_text("text")

        skipped_duplicates = []  # Add this at the top with other init vars

        def save_transaction(tx_data):
            nonlocal bn_counter
            if tx_data and 'amount' in tx_data:
                tx_data['description'] = " ".join(multiline_description).strip() # Join multiline description
                exists = EarmarkedTransaction.objects.filter(
                    date=tx_data['date'],
                    account_name=tx_data['account_name'],
                    amount=tx_data['amount'],
                    property=property_obj
                ).exists()
                if not exists:
                    txn = EarmarkedTransaction(
                        date=tx_data['date'],
                        account_name=tx_data['account_name'],
                        amount=tx_data['amount'],
                        is_income=tx_data['is_income'],
                        description=tx_data['description'],
                        property=property_obj,
                        booking_no=f"{bn_counter:04d}",
                    )
                    earmarked_transactions.append(txn) # Append to earmarked transactions
                    bn_counter += 1 
                    return True 
                else:
                    skipped_duplicates.append(tx_data.copy())
            return False


        pages = split_pdf_into_pages(statement)
        for page_stream in pages:
            text = extract_text_with_fallback(page_stream)
            if not text:
                continue
            lines = text.split('\n')
            current_transaction = {}
            multiline_description = []

            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if "Buchungsdatum:" in line: # Check for booking date
                    match = re.search(r"\d{2}\.\d{2}\.\d{4}", line) # Match date format
                    if match:
                        try:
                            current_buchungsdatum = match.group(0)
                            formatted_buchungsdatum = datetime.strptime(current_buchungsdatum, "%d.%m.%Y").date()
                        except ValueError:
                            formatted_buchungsdatum = None
                    continue

                if any(skip in line for skip in ["Kontostand", "Alter Kontostand", "Neuer Kontostand", "Freistellungsauftrag"]):
                    continue

                if current_buchungsdatum:
                    date_match = re.search(r"\d{2}\.\d{2}", line)
                    if date_match:
                        try:
                            parsed_date = date_match.group(0)
                            formatted_parsed_date = datetime.strptime(parsed_date, "%d.%m").date()
                            formatted_parsed_date = formatted_parsed_date.replace(year=formatted_buchungsdatum.year)
                            if formatted_parsed_date != formatted_buchungsdatum:
                                continue
                        except ValueError:
                            continue

                        save_transaction(current_transaction)

                        current_transaction = {'date': formatted_parsed_date}
                        multiline_description = []

                        current_transaction['account_name'] = line.split(parsed_date)[0].strip()
                        for col in re.split(r'\s{2,}', line):
                            amount_match = re.search(r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?", col)
                            if amount_match:
                                amt = amount_match.group(0).replace('.', '').replace(',', '.')
                                current_transaction['amount'] = abs(float(amt.strip('-')))
                                current_transaction['is_income'] = not col.strip().endswith('-')
                                break
                    elif current_transaction:
                        multiline_description.append(line)

            save_transaction(current_transaction)
            print(f"Processed page {page_stream}: {len(earmarked_transactions)} transactions found.")

        if duplicate_count > 0:
            messages.warning(request, "Einige Transaktionen wurden möglicherweise übersprungen, weil sie bereits vorhanden sind.")

        for txn in earmarked_transactions:
            txn.save()


        request.session['skipped_duplicates'] = serialize_for_session(skipped_duplicates)

        return redirect(f"{reverse('property_detail', args=[property_id])}?tab=dashboard")

#################################################################

# All views for Properties
# Property List View
@login_required
def properties(request):
    query = request.GET.get('q', '')
    
    if query:
        properties = Property.objects.filter(name__icontains=query).order_by('-id')
    else:
        properties = Property.objects.all().order_by('-id')  # Newest first by ID
    
    paginator = Paginator(properties, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    landlords = Landlord.objects.all()
    return render(request, 'bookkeeping/properties.html', {
        'page_obj': page_obj,
        'landlords': landlords,
    })

# Add Property
def add_property(request):
    if request.method == 'POST':
        # Get property fields
        property_type = request.POST.get('property_type')
        name = request.POST.get('name')
        street = request.POST.get('street')
        building_no = request.POST.get('building_no')
        city = request.POST.get('city')
        zip_code = request.POST.get('zip')
        country = request.POST.get('country')
        landlord_ids = request.POST.getlist('landlords')
        # partial_tax_rate = request.POST.get('partial_tax_rate')
        image = request.FILES.get('image')

        auto_calculate_tax = request.POST.get('auto_calculate_tax') == "on"  # Checkbox returns "on" if checked
        
        tax_calculation_method = request.POST.get('tax_calculation_method', 'none')  # Default to manual entry
        
        partial_tax_rate_input = request.POST.get('partial_tax_rate', '0').replace(',', '.')
        partial_tax_rate = float(partial_tax_rate_input) if not auto_calculate_tax else 0.0

        property_obj = Property.objects.create(
            property_type=property_type,
            name=name,
            street=street,
            building_no=building_no,
            city=city,
            zip=zip_code,
            country=country,
            partial_tax_rate=partial_tax_rate,
            auto_calculate_tax=auto_calculate_tax,
            tax_calculation_method=tax_calculation_method,
            image=image,
        )

        property_obj.landlords.set(landlord_ids)  # Assign landlords

        return redirect('properties')

    landlords = Landlord.objects.all()
    return render(request, 'bookkeeping/add_property.html', {
        'landlords': landlords,
    })

# Edit Property
def edit_property(request, pk):
    property_obj = get_object_or_404(Property, id=pk)

    if request.method == 'POST':
        # Update property fields
        property_obj.property_type = request.POST.get('property_type')
        property_obj.name = request.POST.get('name')
        property_obj.street = request.POST.get('street')
        property_obj.building_no = request.POST.get('building_no')
        property_obj.city = request.POST.get('city')
        property_obj.zip = request.POST.get('zip')
        property_obj.country = request.POST.get('country')

        # Convert German-formatted numbers
        auto_calculate_tax = request.POST.get('auto_calculate_tax') == "on"
        print('auto',auto_calculate_tax)
        tax_calculation_method = request.POST.get('tax_calculation_method', 'none')
        print('method',tax_calculation_method)

        partial_tax_rate_input = request.POST.get('partial_tax_rate', '0').replace(',', '.')
        partial_tax_rate = float(partial_tax_rate_input) if not auto_calculate_tax else None

        # Auto-calculation logic
        if auto_calculate_tax:
            if tax_calculation_method == 'sq_meterage':
                leases = Lease.objects.filter(property=property_obj)
                area_with_ust = sum(lease.unit.floor_area for lease in leases if lease.ust_type == 'Mit' and lease.unit.floor_area)
                total_area = sum(lease.unit.floor_area for lease in leases if lease.unit.floor_area)
                partial_tax_rate = round((area_with_ust / total_area) * 100, 2) if total_area > 0 else 0.0

            elif tax_calculation_method == 'income':
                leases = Lease.objects.filter(property=property_obj)
                income_with_ust = sum(
                    (lease.rent + (lease.additional_costs or 0) + lease.deposit_amount)
                    for lease in leases if lease.ust_type == 'Mit'
                )
                total_income = sum(
                    (lease.rent + (lease.additional_costs or 0) + lease.deposit_amount)
                    for lease in leases
                )
                partial_tax_rate = round((income_with_ust / total_income) * 100, 2) if total_income > 0 else 0.0

        # Update the property object with new values
        property_obj.auto_calculate_tax = auto_calculate_tax
        property_obj.tax_calculation_method = tax_calculation_method
        property_obj.partial_tax_rate = partial_tax_rate

        # Update landlords
        landlord_ids = request.POST.getlist('landlords')
        property_obj.landlords.set(landlord_ids)

        # Save and release lock
        property_obj.locked_by = None  # FIXED: was `property.locked_by`
        property_obj.save()

        return redirect('property_detail', property_id=pk)

    landlords = Landlord.objects.all()
    return render(request, 'bookkeeping/edit_property.html', {
        'property': property_obj,
        'landlords': landlords,
    })

# Delete Property
def delete_property(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)

    if request.method == 'POST':
        property_obj.delete()
        return redirect('properties')

    return render(request, 'bookkeeping/delete_property.html', {'property': property_obj})

#################################################################

# All views for Tenants
# Tenant List View
@login_required
def tenants(request):
    query = request.GET.get('q', '')
    property_id = request.GET.get('property')

    tenants = Tenant.objects.all()
    # Filter tenants based on query and property
    if query:
        tenants = tenants.filter(name__icontains=query)
    # Filter by property if provided
    if property_id:
        tenants = tenants.filter(leases__property_id=property_id).distinct()

    # Annotate each tenant with total monthly debt (rent + additional costs)
    tenants = tenants.annotate(
        total_monthly_debt=Sum(
            ExpressionWrapper(
                F('leases__rent') + F('leases__additional_costs'),
                output_field=FloatField()
            )
        )
    )
    # Get all properties for the dropdown
    properties = Property.objects.all()
    
    return render(request, 'bookkeeping/tenants.html', {
        'tenants': tenants,
        'properties': properties,
        'selected_property_id': property_id,
    })

# Add Tenant
@login_required
def add_tenant(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        other_account_names = request.POST.get('other_account_names')
        phone_number = request.POST.get('phone_number')
        email = request.POST.get('email')
        address = request.POST.get('address')
        iban = request.POST.get('iban')
        bic = request.POST.get('bic')
        notes = request.POST.get('notes')

        # Create the tenant
        Tenant.objects.create(
            name=name,
            other_account_names=other_account_names,
            phone_number=phone_number,
            email=email,
            address=address,
            iban=iban,
            bic=bic,
            notes=notes,
        )
        return redirect('tenants')

    return render(request, 'bookkeeping/add_tenant.html')

# Edit Tenant
def edit_tenant(request, pk):
    tenant = get_object_or_404(Tenant, id=pk) # Fetch the tenant using the primary key

    if request.method == 'POST':
        tenant.name = request.POST.get('name')
        tenant.other_account_names = request.POST.get('other_account_names')
        tenant.phone_number = request.POST.get('phone_number')
        tenant.email = request.POST.get('email')
        tenant.address = request.POST.get('address')
        tenant.iban = request.POST.get('iban')
        tenant.bic = request.POST.get('bic')
        tenant.notes = request.POST.get('notes')

        tenant.save()
        return redirect('tenants')

    return render(request, 'bookkeeping/edit_tenant.html', {'tenant': tenant})

# Delete Tenant
def delete_tenant(request, pk):
    tenant = get_object_or_404(Tenant, id=pk)

    if request.method == 'POST':
        tenant.delete()
        return redirect('tenants')

    return render(request, 'bookkeeping/delete_tenant.html', {'tenant': tenant})

#################################################################

@login_required
def mapping_rules(request, property_id):
    property_obj = get_object_or_404(Property, id=property_id) # Get the property object

    expense_profiles = ExpenseProfile.objects.filter(property=property_obj) # Fetch expense profiles for the property
    income_profiles = IncomeProfile.objects.filter(property=property_obj) # Fetch income profiles for the property

    # Merge and sort by ID (or `date` if added later)
    merged_profiles = sorted(
        chain(expense_profiles, income_profiles),
        key=attrgetter('id')  # Or use 'date' if consistent
    )

    return render(request, 'bookkeeping/mapping_rules.html', {
        'property': property_obj,
        'profiles': merged_profiles
    })

#################################################################

# Expense Profiles
def expense_profiles(request):
    properties = Property.objects.prefetch_related('leases', 'expense_profiles').all()
    return render(request, 'bookkeeping/expense_profiles.html', {
        'properties': properties,
    })

# Add Expense Profile
def add_expense_profile(request):
    if request.method == 'POST':
        data = request.POST # Get the POST data 
        invoice_file = request.FILES.get('invoice') # Get the uploaded file object
        invoice_url = None # Initialize invoice URL

        print("POST data received:", data)
        print("Uploaded file object:", invoice_file)

        # Validate Property
        property_id = data.get('property')
        if not property_id:
            print("Missing property ID.")
            return redirect('dashboard')

        property_obj = get_object_or_404(Property, id=property_id)
        print("Found property:", property_obj)

        # Optional Lease
        lease_id = data.get('lease')
        lease = Lease.objects.filter(id=lease_id).first() if lease_id else None
        if lease:
            print("Found lease:", lease)

        # Safe parsing
        amount = safe_decimal(data.get('amount'))
        date = safe_date(data.get('date'))
        recurring = data.get('recurring') == 'on'
        frequency = data.get('frequency')
        ust = int(data.get('ust') or 19)

        print("Parsed amount:", amount)
        print("Parsed date:", date)
        print("Parsed ust:", ust)
        print("Recurring:", recurring)
        print("Frequency:", frequency)

        # Required fields check
        if not data.get('transaction_type') or not data.get('account_name'):
            print("Missing transaction_type or account_name.")
            return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=mapping_tab")

        # Upload to OneDrive
        if invoice_file:
            try:
                print("Attempting OneDrive upload for:", invoice_file.name)
                token = request.session.get('onedrive_token')['access_token']
                if not token:
                    print("No OneDrive token found in session.")
                print("Token retrieved:", token)
                filename = invoice_file.name
                file_content = invoice_file.read()
                print("File content read successfully.")
                invoice_url = upload_to_onedrive(token, file_content, filename)
                print("OneDrive upload successful. URL:", invoice_url)
            except Exception as e:
                print("OneDrive upload failed:", str(e))

        try:
            profile = ExpenseProfile.objects.create(
                lease=lease,
                property=property_obj,
                transaction_type=data.get('transaction_type'),
                account_name=data.get('account_name'),
                amount=amount,
                date=date,
                recurring=recurring,
                frequency=frequency,
                ust=ust,
                invoice=invoice_url
            )
            print("ExpenseProfile created:", profile)
        except ValidationError as e:
            print("Validation error during ExpenseProfile creation:", e)

        return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=mapping_tab")

    print("Request method was not POST.")
    return redirect('dashboard')


# Edit Expense Profile
def edit_expense_profile(request, pk):
    expense = get_object_or_404(ExpenseProfile, id=pk) # Fetch the expense profile using the primary key
    property_obj = expense.property # Get the property object

    if request.method == 'POST':
        data = request.POST # Get the POST data
        lease_id = data.get('lease') # Get the lease ID from the form data
        amount = data.get('amount') # Get the amount from the form data
        date = data.get('date') # Get the date from the form data
        recurring = data.get('recurring') == 'on' # Check if the recurring checkbox is checked

        # Fetch Lease if provided
        lease = Lease.objects.filter(id=lease_id).first() if lease_id else None

        # Determine UST dynamically
        ust = get_ust_from_lease_or_property(lease, property_obj)

        # Update Expense Profile fields
        expense.lease = lease
        # expense.profile_name = data.get('profile_name')
        expense.transaction_type = data.get('transaction_type')
        expense.recurring = recurring
        expense.frequency = data.get('frequency') if recurring else None
        expense.account_name = data.get('account_name')
        expense.ust = ust
        # expense.booking_no = data.get('booking_no')
        
        # Safely parse optional fields
        expense.amount = safe_decimal(amount)
        expense.date = safe_date(date)

        expense.save()
        return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=mapping_tab")

    return redirect('dashboard')  # Default fallback

# Helper function to determine UST based on lease or property
def get_ust_from_lease_or_property(lease, property_obj):
    """
    Determines UST based on lease if available, otherwise defaults to the property UST.
    """
    ust_mapping = {'Voll': 19, 'Teilw': 7, 'Nicht': 0} 
    if lease and lease.ust_type:
        return ust_mapping.get(lease.ust_type, 0)
    return ust_mapping.get(property_obj.ust_type, 0)

# Helper function to safely convert values to Decimal and date
def safe_decimal(value):
    """
    Safely converts a value to Decimal. Returns None if conversion fails.
    """
    try:
        return Decimal(value) if value else None
    except (ValueError, TypeError, Decimal.InvalidOperation):
        return None

# Helper function to safely parse date strings
def safe_date(value):
    """
    Safely parses a date string into a datetime object. Returns None if parsing fails.
    """
    try:
        return datetime.strptime(value, "%Y-%m-%d").date() if value else None
    except (ValueError, TypeError):
        return None
        
# Delete Expense Profile
def delete_expense_profile(request, pk):
    expense_profile = get_object_or_404(ExpenseProfile, id=pk)
    property_obj = expense_profile.property

    if request.method == 'POST':
        expense_profile.delete()
        return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=mapping_tab")

    return redirect('dashboard')  # Default fallback

#################################################################

# Units
def units(request):
    units = Unit.objects.select_related('property').all() # Fetch all units with related property data
    properties = Property.objects.all() # Fetch all properties for the dropdown

    return render(request, 'bookkeeping/units.html',
     {'units': units,
        'properties': properties
     })

# Add Unit
def add_unit(request):
    if request.method == 'POST':
        property_id = request.POST.get('property')
        unit_name = request.POST.get('unit_name')
        floor_area = request.POST.get('floor_area')
        rent = request.POST.get('rent')
        additional_costs = request.POST.get('additional_costs')
        floor = request.POST.get('floor')
        stellplatz_miete = request.POST.get('stellplatz_miete')
        investition_miete = request.POST.get('investition_miete')
        print("POST data received:", request.POST)

        # Ensure the property exists
        property_obj = get_object_or_404(Property, id=property_id)

        # Convert data to appropriate types
        floor_area = float(floor_area) if floor_area else None
        rent = Decimal(rent) if rent else None
        additional_costs = Decimal(additional_costs) if additional_costs else None
        floor = int(floor) if floor else None
        stellplatz_miete = Decimal(stellplatz_miete) if stellplatz_miete else None
        investition_miete = Decimal(investition_miete) if investition_miete else None

        # Create the unit
        Unit.objects.create(
            property=property_obj,
            unit_name=unit_name,
            floor_area=floor_area,
            rent=rent,
            additional_costs=additional_costs,
            floor=floor,
            stellplatz_miete=stellplatz_miete,
            investition_miete=investition_miete
        )

        # Redirect to property detail view
        return redirect('property_detail', property_id=property_obj.id)

    # Fetch properties for the dropdown
    properties = Property.objects.all()
    return render(request, 'bookkeeping/add_unit.html', {'properties': properties})

# Edit Unit
def edit_unit(request, pk):
    unit = get_object_or_404(Unit, id=pk)  # Fetch the unit using the primary key

    if request.method == 'POST':
        # Update unit details from the form data
        unit.unit_name = request.POST.get('unit_name')
        unit.floor_area = request.POST.get('floor_area')
        unit.rent = request.POST.get('rent')
        unit.additional_costs = request.POST.get('additional_costs')
        unit.floor = request.POST.get('floor')
        unit.position = request.POST.get('position')
        unit.stellplatz_miete = request.POST.get('stellplatz_miete')
        unit.investition_miete = request.POST.get('investition_miete')

        # Convert data to appropriate types
        unit.floor_area = float(unit.floor_area) if unit.floor_area else None
        unit.rent = Decimal(unit.rent) if unit.rent else None
        unit.additional_costs = Decimal(unit.additional_costs) if unit.additional_costs else None
        unit.floor = int(unit.floor) if unit.floor else None
        unit.stellplatz_miete = Decimal(unit.stellplatz_miete) if unit.stellplatz_miete else None
        unit.investition_miete = Decimal(unit.investition_miete) if unit.investition_miete else None

        # Retrieve the property using the unit's property relationship
        property_id = unit.property.id  # Fetch the associated property ID
        unit.property = get_object_or_404(Property, id=property_id)

        # Save the updated unit
        unit.save()

        # Redirect to the associated property's detail view
        return redirect('property_detail', property_id=property_id)

    # Fetch all properties for the dropdown (if needed, though it's redundant in this case)
    properties = Property.objects.all()
    return render(request, 'bookkeeping/edit_unit.html', {
        'unit': unit,
        'properties': properties,
    })

# Delete Unit
def delete_unit(request, pk):
    unit = get_object_or_404(Unit, id=pk)
    property_id = unit.property.id  # Store the property ID before deletion

    if request.method == 'POST':
        unit.delete()
        # Redirect to property detail view after deleting the unit
        return redirect('property_detail', property_id=property_id)

    return render(request, 'bookkeeping/delete_unit.html', {'unit': unit})

#################################################################

# Landlords View
def landlords(request):
    query = request.GET.get('q', '') # Search query from the GET request 
    property_id = request.GET.get('property') # Property ID from the GET request

    landlords = Landlord.objects.prefetch_related('owned_properties').all() # Fetch all landlords with related properties

    if query:
        landlords = landlords.filter(name__icontains=query) # Filter landlords by name

    if property_id:
        landlords = landlords.filter(owned_properties__id=property_id).distinct() # Filter landlords by property ID
    properties = Property.objects.all() # Fetch all properties for the dropdown

    return render(request, 'bookkeeping/landlords.html', {
        'landlords': landlords,
        'properties': properties,
        'selected_property_id': property_id,
    })

# Add Landlord
def add_landlord(request):
    if request.method == 'POST':
        
        name = request.POST.get('name')
        phone_number = request.POST.get('phone_number')
        email = request.POST.get('email')
        address = request.POST.get('address')
        iban = request.POST.get('iban')
        bic = request.POST.get('bic')
        tax_id = request.POST.get('tax_id')
        objects  = request.POST.get('object')
        notes = request.POST.get('notes')

        # Create the landlord
        Landlord.objects.create(
            name=name,
            phone_number=phone_number,
            email=email,
            address=address,
            iban=iban,
            bic=bic,
            tax_id=tax_id,
            object =objects ,
            notes=notes,
        )
        return redirect('landlords')

    return render(request, 'bookkeeping/add_landlord.html')

# Edit Landlord
def edit_landlord(request, pk):
    landlord = get_object_or_404(Landlord, id=pk) # Fetch the landlord using the primary key

    if request.method == 'POST':
        landlord.name = request.POST.get('name')
        landlord.phone_number = request.POST.get('phone_number')
        landlord.email = request.POST.get('email')
        landlord.address = request.POST.get('address')
        landlord.iban = request.POST.get('iban')
        landlord.bic = request.POST.get('bic')
        landlord.tax_id = request.POST.get('tax_id')
        landlord.object  = request.POST.get('object ')
        landlord.notes = request.POST.get('notes')

        landlord.save()
        return redirect('landlords')

    return render(request, 'bookkeeping/edit_landlord.html', {'landlord': landlord})

# Delete Landlord
def delete_landlord(request, pk):
    landlord = get_object_or_404(Landlord, id=pk) # Fetch the landlord using the primary key

    if request.method == 'POST':
        landlord.delete()
        return redirect('landlords')

    return render(request, 'bookkeeping/delete_landlord.html', {'landlord': landlord})

#################################################################

# Export parsed transactions to Excel

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
from django.http import HttpResponse
from django.utils.timezone import now
from django.utils.text import slugify
import calendar

def export_parsed_transactions(request, property_id):


    german_months = [
        "", "Januar", "Februar", "März", "April", "Mai", "Juni",
        "Juli", "August", "September", "Oktober", "November", "Dezember"
    ]

    property_obj = Property.objects.get(id=property_id)

    # ✅ Sort by booking_no
    parsed_transactions = ParsedTransaction.objects.filter(
        related_property=property_obj
    ).order_by('booking_no')

    wb = Workbook()
    ws = wb.active
    ws.title = "Parsed Transactions"

    headers = [
        "BN", "Datum", "Kontoname", "Einheit",
        "Bank", "Nettobetrag", "Betrag Aufw", "USt (%)", "USt",
        "Bruttobetrag", "Gggkto", "Mieter", "Rechnung",
        "Monat", "L-Monat", "Objekt"
    ]
    ws.append(headers)

    for tx in parsed_transactions:
        month_int = tx.date.month if tx.date else None
        month_name = german_months[month_int] if month_int else "N/A"

        # ✅ Safely calculate VAT
        calculated_ust = None
        if tx.betrag_brutto and tx.ust_type:
            try:
                calculated_ust = round(tx.betrag_brutto - (tx.betrag_brutto / (1 + tx.ust_type / 100)), 2)
            except:
                calculated_ust = None

        ws.append([
            tx.booking_no or "N/A",
            tx.date.strftime("%Y-%m-%d") if tx.date else "N/A",
            tx.account_name or "N/A",
            tx.unit_name or "N/A",

            tx.betrag_brutto if tx.betrag_brutto is not None else None,  # Bank
            tx.betrag_netto if tx.betrag_netto is not None else None,    # Nettobetrag
            tx.betrag_netto if tx.betrag_netto is not None else None,    # Betrag Aufw

            f"{tx.ust_type}%" if tx.ust_type is not None else "N/A",     # USt %
            calculated_ust,                                              # Corrected USt

            tx.betrag_brutto if tx.betrag_brutto is not None else None,  # Bruttobetrag
            tx.transaction_type or "N/A",                                # Gggkto
            tx.tenant or "N/A",                                          # Mieter
            tx.invoice.name if tx.invoice else "-",                      # Rechnung

            month_int or "N/A",
            month_name,
            property_obj.name
        ])

    # Format numeric cells
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=10):  # Bank to Bruttobetrag
        for cell in row:
            if isinstance(cell.value, (float, int)):
                cell.number_format = numbers.FORMAT_NUMBER_00

    # Excel table formatting
    def col_letter(n):
        result = ""
        while n:
            n, rem = divmod(n - 1, 26)
            result = chr(65 + rem) + result
        return result

    num_rows = ws.max_row
    num_cols = ws.max_column
    table_ref = f"A1:{col_letter(num_cols)}{num_rows}"

    table = Table(displayName="ParsedTransactionsTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(cell_value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    today = now().strftime("%Y-%m-%d")
    property_name_slug = slugify(property_obj.name)
    filename = f"FiBu_{property_name_slug}_{today}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



#################################################################

# leases
def leases(request):
    leases = Lease.objects.select_related('property', 'unit', 'tenant').prefetch_related('landlords').all()
    properties = Property.objects.all()
    units = Unit.objects.all()
    tenants = Tenant.objects.all()
    landlords = Landlord.objects.all()

    return render(request, 'bookkeeping/leases.html', {
        'leases': leases,
        'properties': properties,
        'units': units,
        'tenants': tenants,
        'landlords': landlords,
    })

# Add Lease
def add_lease(request):
    if request.method == 'POST':
        unit_id = request.POST.get('unit')
        tenant_ids = request.POST.getlist('tenants')  # Get multiple tenants as a list
        start_date = request.POST.get('start_date') or None
        end_date = request.POST.get('end_date') or None
        ust_type = request.POST.get('ust_type')
        deposit_amount = request.POST.get('deposit_amount') or Decimal(0)
        guarantee = request.POST.get('guarantee') or Decimal(0)
        # Convert deposit amount safely
        deposit_amount = Decimal(deposit_amount.replace(',', '.')) if deposit_amount else Decimal(0)
        guarantee = Decimal(guarantee.replace(',', '.')) if guarantee else Decimal(0)
        print("POST data received:", request.POST)
        # Fetch the unit and its associated property, landlords, and rent
        unit_obj = get_object_or_404(Unit, id=unit_id)
        property_obj = unit_obj.property
        landlords = property_obj.landlords.all()
        rent = unit_obj.rent
        
        # Fetch tenants and extract data
        tenants = Tenant.objects.filter(id__in=tenant_ids)
        account_names = [tenant.name for tenant in tenants]
        ibans = [tenant.iban for tenant in tenants]

        additional_costs = unit_obj.additional_costs  # Fetch additional costs from the unit

        lease = Lease.objects.create(
            property=property_obj,
            unit=unit_obj,
            start_date=start_date,
            end_date=end_date,
            ust_type=ust_type,
            deposit_amount=deposit_amount,
            rent=rent,
            additional_costs=additional_costs,  # Store in lease
            account_names=account_names,
            ibans=ibans,
            guarantee=guarantee, 
        )

        lease.landlords.set(landlords)  # Set landlords for the lease
        lease.tenants.set(tenants)  # Assign multiple tenants
        lease.save()

        # Redirect to property detail view
        return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=leases")
    
    # Fetch data for dropdowns if needed
    properties = Property.objects.all()
    units = Unit.objects.all()
    tenants = Tenant.objects.all()
    return render(request, 'bookkeeping/add_lease.html', {
        'properties': properties,
        'units': units,
        'tenants': tenants,
    })

# Edit Lease
def edit_lease(request, pk):
    lease = get_object_or_404(Lease, id=pk)

    if request.method == 'POST':
        lease.unit = get_object_or_404(Unit, id=request.POST.get('unit'))
        lease.start_date = request.POST.get('start_date') or None
        lease.end_date = request.POST.get('end_date') or None
        lease.ust_type = request.POST.get('ust_type')

        # Convert deposit amount safely
        deposit_amount = request.POST.get('deposit_amount', '0').replace(',', '.')
        lease.deposit_amount = Decimal(deposit_amount)

        # Fetch unit's rent
        lease.rent = lease.unit.rent

        lease.additional_costs = lease.unit.additional_costs  # Update from unit

        # Save updated landlords
        landlord_ids = request.POST.getlist('landlords')
        lease.landlords.set(landlord_ids)

        # Handle multiple tenants
        tenant_ids = request.POST.getlist('tenants')
        tenants = Tenant.objects.filter(id__in=tenant_ids)
        lease.tenants.set(tenants)

        # Update account names & IBANs for tenants
        lease.account_names = [tenant.name for tenant in tenants]
        lease.ibans = [tenant.iban for tenant in tenants]

        # Save the updated lease
        lease.save()

        # Redirect to property detail view
        return redirect(f"{reverse('property_detail', args=[lease.property.id])}?tab=leases")
    
    # Fetch data for dropdowns
    properties = Property.objects.all()
    units = Unit.objects.all()
    tenants = Tenant.objects.all()
    landlords = Landlord.objects.all()

    return render(request, 'bookkeeping/edit_lease.html', {
        'lease': lease,
        'properties': properties,
        'units': units,
        'tenants': tenants,
        'landlords': landlords
    })

# Delete Lease
def delete_lease(request, pk):
    lease = get_object_or_404(Lease, id=pk)
    property_id = lease.property.id  # Store the associated property ID before deletion

    if request.method == 'POST':
        lease.delete() # Delete the lease object 

        # Redirect to property detail view
        return redirect(f"{reverse('property_detail', args=[lease.property.id])}?tab=leases")

    return render(request, 'bookkeeping/delete_lease.html', {'lease': lease})

#################################################################

# Income Profiles
def income_profiles(request):
    properties = Property.objects.prefetch_related('leases__income_profiles', 'income_profiles')
    return render(request, 'bookkeeping/income_profiles.html', {'properties': properties})

# Add Income Profile
def add_income_profile(request):
    if request.method == 'POST':
        lease_ids = request.POST.getlist('leases')
        property_id = request.POST.get('property')
        date = request.POST.get('date')
        ust = int(request.POST.get('ust', 0))
        recurring = request.POST.get('recurring') == 'on'
        frequency = request.POST.get('frequency')
        account_name = request.POST.get('account_name')

        leases = Lease.objects.filter(id__in=lease_ids)
        property_obj = get_object_or_404(Property, id=property_id)

        try:
            date = datetime.strptime(date, "%Y-%m-%d").date() if date else None
        except Exception:
            date = None

        try:
            total_amount = Decimal(request.POST.get('amount', '0'))
        except Exception:
            total_amount = Decimal("0.00")

        split_enabled = request.POST.get("split_income") == "on"
        apply_remainder = request.POST.get("apply_remainder") == "on"

        if split_enabled:
            selected_types = request.POST.getlist("split_types[]")
            split_total = Decimal("0.00")
            split_map = {}

            for tx_type in selected_types:
                field_key = f"split_amount_{tx_type}"
                try:
                    split_amount = Decimal(request.POST.get(field_key, "0"))
                except Exception:
                    split_amount = Decimal("0.00")

                if split_amount > 0:
                    split_map[tx_type] = float(split_amount)
                    split_total += split_amount

            remainder = total_amount - split_total

            if remainder > 0 and apply_remainder:
                split_map["account_balance"] = float(remainder)

            income_profile = IncomeProfile.objects.create(
                property=property_obj,
                transaction_type=None,
                amount=total_amount,
                date=date,
                recurring=recurring,
                frequency=frequency,
                account_name=account_name,
                ust=ust,
                split_details=split_map
            )

            income_profile.leases.set(leases)

            # Now apply remainder AFTER leases are saved
            if remainder > 0 and apply_remainder:
                lease = leases.first()
                if lease and lease.tenants.exists():
                    tenant = lease.tenants.first()
                    tenant.balance += remainder
                    tenant.save()
                    messages.info(request, f"Restbetrag von €{remainder:.2f} wurde dem Saldo des Mieters gutgeschrieben.")
            elif remainder > 0:
                messages.info(request, f"Restbetrag von €{remainder:.2f} wurde nicht zugewiesen.")

        else:
            income_profile = IncomeProfile.objects.create(
                property=property_obj,
                transaction_type=request.POST.get('transaction_type'),
                amount=total_amount,
                date=date,
                recurring=recurring,
                frequency=frequency,
                account_name=account_name,
                ust=ust,
                split_details=None
            )
            income_profile.leases.set(leases)

        return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=mapping_tab")

    return redirect('dashboard')

from django.http import JsonResponse
from .models import Lease

def matching_leases(request):
    account_name = request.GET.get('account_name', '')
    leases = Lease.objects.filter(account_names__icontains=account_name)
    lease_ids = list(leases.values_list('id', flat=True))
    return JsonResponse({'lease_ids': lease_ids})


# Edit Income Profile
def edit_income_profile(request, pk):
    income = get_object_or_404(IncomeProfile, id=pk) # Fetch the income profile using the primary key
    property_obj = income.property # Get the property object

    if request.method == 'POST':
        lease_id = request.POST.get('lease')
        property_id = request.POST.get('property')
        amount = request.POST.get('amount')
        date = request.POST.get('date')

        # Fetch Lease and Property
        lease = Lease.objects.filter(id=lease_id).first() if lease_id else None

        # Determine UST dynamically
        ust = 0
        if lease and lease.ust_type:
            ust = {'Voll': 19, 'Teilw': 7, 'Nicht': 0}.get(lease.ust_type, 0)

        # Safely parse optional fields
        try:
            income.amount = Decimal(amount) if amount else None
        except Exception:
            income.amount = None
        try:
            income.date = datetime.strptime(date, "%Y-%m-%d").date() if date else None
        except Exception:
            income.date = None

        # Update fields
        income.lease = lease
        income.property = property_obj
        # income.profile_name = request.POST.get('profile_name')
        income.transaction_type = request.POST.get('transaction_type')
        income.recurring = request.POST.get('recurring') == 'on'
        income.frequency = request.POST.get('frequency') if income.recurring else None
        income.account_name = request.POST.get('account_name')
        income.ust = ust
        # income.booking_no = request.POST.get('booking_no')

        income.save()
        return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=mapping_tab")

    return redirect('dashboard')  # Default fallback


# Delete Income Profile
def delete_income_profile(request, pk):
    income_profile = get_object_or_404(IncomeProfile, id=pk)
    property_obj = income_profile.property

    if request.method == 'POST':
        income_profile.delete()
        return redirect(f"{reverse('property_detail', args=[property_obj.id])}?tab=mapping_tab")

    return redirect('dashboard')  # Default fallback

#################################################################

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Unit, Tenant

def fetch_unit_tenant_data(request):
    unit_id = request.GET.get('unit_id')
    tenant_ids = request.GET.getlist('tenant_ids[]')  # Get multiple tenant IDs

    response = {}

    if unit_id:
        unit = get_object_or_404(Unit, id=unit_id)
        response['property_id'] = unit.property.id
        response['landlord_ids'] = list(unit.property.landlords.values_list('id', flat=True))
        response['rent'] = unit.rent

    if tenant_ids:
        tenants = Tenant.objects.filter(id__in=tenant_ids)
        response['account_names'] = [tenant.name for tenant in tenants]
        response['ibans'] = [tenant.iban for tenant in tenants]

    return JsonResponse(response)

#################################################################

# AJAX endpoint to fetch lease profiles
from django.http import JsonResponse
from django.http import JsonResponse
from decimal import Decimal

def lease_profiles(request, lease_id):
    lease = get_object_or_404(Lease, id=lease_id) # Fetch the lease object
    # Check if the user is authenticated
    income_data = [] # Initialize income data list
    # Rent
    if lease.rent:
        income_data.append({
            "type": "rent",
            "label": "Miete",
            "amount": float(lease.rent)
        })
    # BK / Additional Costs
    additional = lease.additional_costs or (lease.unit.additional_costs if lease.unit else Decimal("0.00"))
    if additional:
        income_data.append({
            "type": "bk_advance_payments",
            "label": "Nebenkosten",
            "amount": float(additional)
        })
    # Deposit
    if lease.deposit_amount:
        income_data.append({
            "type": "security_deposit",
            "label": "Kaution",
            "amount": float(lease.deposit_amount)
        })
    return JsonResponse({
        "lease_id": lease.id,
         "unit_name": lease.unit.unit_name if lease.unit else None,
        "incomes": income_data
    })

from django.db.models import Sum, Avg
from django.utils.timezone import now, timedelta
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import Property

def get_property_locks(request):
    if request.user.is_authenticated:
        locks = Property.objects.filter(locked_by__isnull=False).values('id', 'locked_by__username')
        return JsonResponse(list(locks), safe=False)
    return JsonResponse({"error": "Unauthorized"}, status=403)


@csrf_exempt
def unlock_property(request, property_id):
    if request.method == "POST" and request.user.is_authenticated:
        try:
            prop = Property.objects.get(id=property_id)
            if prop.locked_by == request.user:
                prop.locked_by = None
                prop.save(update_fields=["locked_by"])
                print(f"[unlock_property] 🔓 Unlocked property {property_id} by {request.user.username}")
                return JsonResponse({"status": "unlocked"})
        except Property.DoesNotExist:
            pass
    return JsonResponse({"status": "error"}, status=400)

def property_detail(request, property_id):
    property_obj = get_object_or_404(Property, id=property_id)

    # FIXED: reference was to `property`, changed to `property_obj`
    locked_by_another_user = False
    if property_obj.locked_by and property_obj.locked_by != request.user:
        locked_by_another_user = True
    else:
        if property_obj.locked_by != request.user:
            property_obj.locked_by = request.user
            property_obj.save(update_fields=['locked_by'])

    # Related data
    leases = property_obj.leases.all()
    income_profiles = IncomeProfile.objects.filter(property=property_obj)
    expense_profiles = ExpenseProfile.objects.filter(property=property_obj)
    #sort earmarked transaction by booking no
    earmarked_transactions = EarmarkedTransaction.objects.filter(property=property_obj).order_by('booking_no')
    #Sort parsed transactions by booking no
    parsed_transactions = ParsedTransaction.objects.filter(related_property=property_obj).order_by('booking_no')
    # Combine and sort all mapping rules by creation order (fallback to ID if needed)
    all_mapping_profiles = sorted(
        chain(income_profiles, expense_profiles),
        key=attrgetter('id')  # Use `date` if available later
    )
    landlords = Landlord.objects.all()
    tenants = Tenant.objects.all()
    property_landlords = property_obj.landlords.all()

    # Financial overview
    total_revenue = income_profiles.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expense_profiles.aggregate(total=Sum('amount'))['total'] or 0
    net_income = total_revenue - total_expenses

    # Unit statistics
    total_units = property_obj.units.count()
    leased_units = leases.count()
    occupancy_rate = (leased_units / total_units) * 100 if total_units else 0
    avg_rent = leases.aggregate(average=Avg('rent'))['average'] or 0

    # Chart data
    chart_data = {
        "labels": [],
        "revenue": [],
        "expenses": []
    }
    # Get the last 6 months of data
    # Assuming the date is in the format YYYY-MM-DD
    current_date = now()
    for i in range(6):
        month = (current_date - timedelta(days=30 * i)).strftime("%B %Y")
        revenue = income_profiles.filter(date__month=(current_date - timedelta(days=30 * i)).month).aggregate(total=Sum('amount'))['total'] or 0
        expense = expense_profiles.filter(date__month=(current_date - timedelta(days=30 * i)).month).aggregate(total=Sum('amount'))['total'] or 0
        chart_data["labels"].append(month)
        chart_data["revenue"].append(revenue)
        chart_data["expenses"].append(expense)

    skipped_duplicates = request.session.pop('skipped_duplicates', None)
    
    context = {
        'property': property_obj,
        'units': property_obj.units.all(),
        'leases': leases,
        'income_profiles': income_profiles,
        'expense_profiles': expense_profiles,
        'profiles': all_mapping_profiles,
        'earmarked_transactions': earmarked_transactions,
        'parsed_transactions': parsed_transactions,
        'financial_overview': {
            'total_revenue': total_revenue,
            'total_expenses': total_expenses,
            'net_income': net_income,
        },
        'unit_stats': {
            'total_units': total_units,
            'leased_units': leased_units,
            'occupancy_rate': occupancy_rate,
            'avg_rent': avg_rent,
        },
        "chart_data": chart_data,
        'landlords': landlords,
        'tenants': tenants,
        'property_landlords': property_landlords,
        'locked_by_another_user': locked_by_another_user,
        'skipped_duplicates': skipped_duplicates,
    }
    print(f"[property_detail] Property {property_obj.id} locked by {property_obj.locked_by}")
    return render(request, 'bookkeeping/property_detail.html', context)


from django.db.models import Sum, F, FloatField
from collections import defaultdict
from calendar import month_name
from django.db.models import Sum, F, FloatField, Case, When, Value as V
from django.shortcuts import render, get_object_or_404
from .models import Property, ParsedTransaction, IncomeProfile

def ust_view(request, property_id):
    property_instance = get_object_or_404(Property, id=property_id)

    default_types = ['rent', 'bk_advance_payment']
    selected_types = request.GET.getlist('transaction_types') or default_types

    income_qs = ParsedTransaction.objects.filter(
        related_property_id=property_id,
        is_income=True,
        transaction_type__in=selected_types
    )

    expense_qs = ParsedTransaction.objects.filter(
        related_property_id=property_id,
        is_income=False
    )

    # Annotate income transactions by month/year
    income_data = income_qs.annotate(
        year=F('date__year'),
        month=F('date__month')
    ).values('year', 'month').annotate(
        umzu19_brutto=Sum(
            Case(When(ust_type=19, then=F('betrag_brutto')), default=V(0), output_field=FloatField())
        ),
        umzu7_brutto=Sum(
            Case(When(ust_type=7, then=F('betrag_brutto')), default=V(0), output_field=FloatField())
        ),
        ust_output=Sum(
            Case(
                When(ust_type=19, then=F('betrag_brutto') - (F('betrag_brutto') / 1.19)),
                When(ust_type=7, then=F('betrag_brutto') - (F('betrag_brutto') / 1.07)),
                default=V(0),
                output_field=FloatField()
            )
        )
    )

    # Annotate expense transactions by month/year
    expense_data = expense_qs.annotate(
        year=F('date__year'),
        month=F('date__month')
    ).values('year', 'month').annotate(
        vorsteuer=Sum(
            Case(
                When(ust_type=19, then=F('betrag_brutto') - (F('betrag_brutto') / 1.19)),
                When(ust_type=7, then=F('betrag_brutto') - (F('betrag_brutto') / 1.07)),
                default=V(0),
                output_field=FloatField()
            )
        )
    )

    # Merge both income and expense data
    monthly_map = defaultdict(lambda: {
        'umzu19_brutto': 0,
        'umzu7_brutto': 0,
        'ust_output': 0,
        'vorsteuer': 0,
    })

    # Add income data
    for row in income_data:
        key = (row['year'], row['month'])
        monthly_map[key]['year'] = row['year']
        monthly_map[key]['month'] = row['month']
        monthly_map[key]['umzu19_brutto'] += row['umzu19_brutto'] or 0
        monthly_map[key]['umzu7_brutto'] += row['umzu7_brutto'] or 0
        monthly_map[key]['ust_output'] += row['ust_output'] or 0

    # Add expense data
    for row in expense_data:
        key = (row['year'], row['month'])
        monthly_map[key]['year'] = row['year']
        monthly_map[key]['month'] = row['month']
        monthly_map[key]['vorsteuer'] += row['vorsteuer'] or 0

    # Final list of merged monthly data with saldo and label
    monthly_data = []
    for key in sorted(monthly_map):
        row = monthly_map[key]
        row['saldo_ust'] = row['ust_output'] - row['vorsteuer']
        row['month_label'] = f"{month_name[row['month']]} {row['year']}"
        monthly_data.append(row)

    # Totals
    totals = {
        'total_umzu19_brutto': sum(d['umzu19_brutto'] for d in monthly_data),
        'total_umzu7_brutto': sum(d['umzu7_brutto'] for d in monthly_data),
        'total_ust_output': sum(d['ust_output'] for d in monthly_data),
        'total_vorsteuer': sum(d['vorsteuer'] for d in monthly_data),
        'total_saldo_ust': sum(d['saldo_ust'] for d in monthly_data),
    }

    context = {
        'property': property_instance,
        'monthly_data': monthly_data,
        'totals': totals,
        'total_19_percent': totals['total_umzu19_brutto'] * 0.19 if totals['total_umzu19_brutto'] else 0,
        'all_transaction_types': IncomeProfile.TRANSACTION_TYPES,
        'selected_types': selected_types,
        'default_types': default_types,
    }

    return render(request, 'bookkeeping/ust.html', context)

#################################################################

import requests
import json
from django.http import JsonResponse
from django.shortcuts import redirect
from django.conf import settings
from urllib.parse import urlencode

# Base API URL for Corporate Payments
COMMERZBANK_BASE_URL = "https://api-sandbox.commerzbank.com/corporate-payments-api/1/v1"

# OAuth2 Endpoints
AUTH_BASE_URL = "https://api-sandbox.commerzbank.com/auth/realms/sandbox/protocol/openid-connect"
AUTHORIZE_URL = f"{AUTH_BASE_URL}/auth"
TOKEN_URL = f"{AUTH_BASE_URL}/token"

def authorize_commerzbank(request):
    """Redirects user to Commerzbank OAuth2 login."""

    if not settings.COMMERZBANK_CLIENT_ID or not settings.COMMERZBANK_CLIENT_SECRET:
        return JsonResponse({"error": "Missing API credentials"}, status=500) # Check for credentials


    params = {
        "response_type": "code",
        "client_id": settings.COMMERZBANK_CLIENT_ID,
        "redirect_uri": settings.COMMERZBANK_REDIRECT_URI,
        "state": "secure_random_string"
    }

    auth_url = f"{AUTHORIZE_URL}?{requests.compat.urlencode(params)}" # Construct the authorization URL
    print(f"[authorize_commerzbank] Redirecting to: {auth_url}")

    return redirect(auth_url)

def commerzbank_callback(request):
    """Handles OAuth2 callback and retrieves the access token."""
    print("[commerzbank_callback] Handling OAuth2 callback...")

    code = request.GET.get("code") # Authorization code from the callback
    # property_id = request.session.get("property_id") # Retrieve property ID from session

    if not code:
        return JsonResponse({"error": "No authorization code received."}, status=400) # Handle missing code

    # if not property_id:
    #     return JsonResponse({"error": "Property ID missing."}, status=400)  # Handle missing property ID

    # Exchange authorization code for access token
    data = {
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": settings.COMMERZBANK_REDIRECT_URI,
        "client_id": settings.COMMERZBANK_CLIENT_ID,   # Use secure storage for secrets in production
        "client_secret": settings.COMMERZBANK_CLIENT_SECRET, # Use secure storage for secrets in production
    }

    headers = {"Content-Type": "application/x-www-form-urlencoded"} # Set headers for the request
    response = requests.post(TOKEN_URL, data=data, headers=headers) # Send POST request to get access token

    if response.status_code == 200:
        token_data = response.json() # Parse the response
        access_token = token_data.get("access_token")  # Extract access token
        request.session["access_token"] = access_token # Store in session
        print(f"[commerzbank_callback] Access Token Retrieved: {access_token[:10]}...") # Log only part of the token for security

        # Redirect to property page with success flag
        return redirect(f"/settings/?auth_success=true")

    else:
        return JsonResponse({"error": "Authentication failed."}, status=400)

def fetch_commerzbank_messages(access_token):
    """Fetch available messages (only camt.053 bank statements) from Commerzbank API."""
    print("[fetch_commerzbank_messages] 🔄 Fetching bank statements (camt.053)...")

    messages_url = f"{COMMERZBANK_BASE_URL}/messages?OrderType=C53"  # Only camt.053
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    response = requests.get(messages_url, headers=headers)
    print(f"[fetch_commerzbank_messages] 🏦 API Response Status: {response.status_code}")

    if response.status_code == 200:
        try:
            messages_data = response.json()
            print(f"[fetch_commerzbank_messages] Bank Statements Retrieved: {json.dumps(messages_data, indent=2)}")
            return messages_data
        except json.JSONDecodeError:
            return {"error": "Invalid JSON response from Commerzbank API."}
    else:
        return {"error": f"Failed to retrieve messages. Status: {response.status_code}"}

import requests
from django.http import JsonResponse
from django.conf import settings

COMMERZBANK_BASE_URL = "https://api-sandbox.commerzbank.com/corporate-payments-api/1/v1/bulk-payments"

def get_commerzbank_transactions(request):
    """Fetches C53 bank statement transactions."""
    print("[get_commerzbank_transactions] 🔄 Fetching bank statements...")

    access_token = request.session.get("access_token")
    if not access_token:
        return JsonResponse({"error": "User not authenticated."}, status=401)

    # Step 1: Get the list of available C53 messages
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "ClientProduct": "Test_AppV0.1"
    }

    response = requests.get(f"{COMMERZBANK_BASE_URL}/messages?OrderType=C53", headers=headers)

    if response.status_code != 200:
        return JsonResponse({"error": "Failed to retrieve messages.", "status": response.status_code}, status=response.status_code)

    messages_data = response.json()

    # Check if response is a list
    if isinstance(messages_data, list):
        messages = messages_data  # Assign it directly
    else:
        messages = messages_data.get("messages", [])  # Fallback if API changes

    transactions = [] ## List to store transactions

    # Step 2: Fetch bank statement details for each message ID
    for message in messages:
        message_id = message.get("MessageId")  # Ensure it's a dict before accessing
        if not message_id:
            continue

        statement_url = f"{COMMERZBANK_BASE_URL}/messages/{message_id}"
        statement_response = requests.get(statement_url, headers=headers)

        if statement_response.status_code == 200:
            transactions.append(statement_response.text)  # XML response (camt.053 format)
        else:
            print(f"[get_commerzbank_transactions] Failed to fetch statement {message_id}, Status: {statement_response.status_code}")

        # Step 3: Confirm retrieval
        confirm_url = f"{COMMERZBANK_BASE_URL}/messages/{message_id}"
        confirm_data = {"received": "complete"}
        confirm_response = requests.put(confirm_url, headers=headers, json=confirm_data)

        if confirm_response.status_code != 200:
            print(f"[get_commerzbank_transactions] Failed to confirm retrieval of {message_id}")

    return JsonResponse({"transactions": transactions})


def extract_transactions_from_camt(statement_details):
    """Extract transactions from camt.053 bank statement XML response."""
    try:
        transactions = []
        import xml.etree.ElementTree as ET

        root = ET.fromstring(statement_details)
        ns = {'ns': "urn:iso:std:iso:20022:tech:xsd:camt.053.001.08"}

        # Iterate through each transaction
        for entry in root.findall(".//ns:Stmt/ns:Ntry", ns):
            amount = entry.find("ns:Amt", ns).text
            currency = entry.find("ns:Amt", ns).attrib.get("Ccy", "EUR")
            date = entry.find("ns:BookgDt/ns:Dt", ns).text
            description = entry.find("ns:AddtlNtryInf", ns).text

            transactions.append({
                "date": date,
                "amount": f"{amount} {currency}",
                "description": description
            })

        return transactions
    except Exception as e:
        print(f"[extract_transactions_from_camt] Error parsing camt.053: {e}")
        return []
    

from django.shortcuts import redirect
from django.conf import settings
from requests_oauthlib import OAuth2Session

def onedrive_auth(request):
    print("[onedrive_auth] Redirecting to OneDrive OAuth2 login...")
    print('OneDrive Client ID:', settings.ONEDRIVE_CLIENT_ID)  # Debugging line
    print('OneDrive Redirect URI:', settings.ONEDRIVE_REDIRECT_URI)  # Debugging line
    print('OneDrive Scopes:', settings.ONEDRIVE_SCOPES)  # Debugging line
    
    oauth = OAuth2Session(
        settings.ONEDRIVE_CLIENT_ID,
        redirect_uri=settings.ONEDRIVE_REDIRECT_URI,
        scope=settings.ONEDRIVE_SCOPES,
    )
    auth_url, state = oauth.authorization_url('https://login.microsoftonline.com/common/oauth2/v2.0/authorize')
    request.session['oauth_state'] = state
    return redirect(auth_url)

def onedrive_callback(request):
    oauth = OAuth2Session(
        settings.ONEDRIVE_CLIENT_ID,
        state=request.session['oauth_state'],
        redirect_uri=settings.ONEDRIVE_REDIRECT_URI,
        scope=settings.ONEDRIVE_SCOPES  # explicitly add scopes here
    )

    token = oauth.fetch_token(
        'https://login.microsoftonline.com/common/oauth2/v2.0/token',
        client_secret=settings.ONEDRIVE_CLIENT_SECRET,
        authorization_response=request.build_absolute_uri(),
        include_client_id=True,
        scope=settings.ONEDRIVE_SCOPES  # explicitly include scope here
    )

    request.session['onedrive_token'] = token
    request.session.modified = True

    return redirect('/')


import requests

def upload_to_onedrive(token, filename, file_content):
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/pdf',
    }
    upload_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/Invoices/{filename}:/content'

    response = requests.put(upload_url, headers=headers, data=file_content)

    if response.status_code in [200, 201]:
        json_response = response.json()
        file_web_url = json_response.get('webUrl')
        return file_web_url  # Save this link to your database
    else:
        raise Exception(f'Failed to upload: {response.content}')


#Settings page view
def settings_view(request):
    if request.method == 'POST':
        # Handle form submission
        # Save settings to the database or session
        pass  # Implement your logic here

    # Render the settings page template
    return render(request, 'bookkeeping/settings.html')